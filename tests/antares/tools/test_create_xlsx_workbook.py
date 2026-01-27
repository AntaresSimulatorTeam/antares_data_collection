# Copyright (c) 2024, RTE (https://www.rte-france.com)
#
# See AUTHORS.txt
#
# This Source Code Form is subject to the terms of the Mozilla Public
# License, v. 2.0. If a copy of the MPL was not distributed with this
# file, You can obtain one at http://mozilla.org/MPL/2.0/.
#
# SPDX-License-Identifier: MPL-2.0
#
# This file is part of the Antares project.

import os
import re
import pytest
from pathlib import Path

import pandas as pd
from openpyxl.reader.excel import load_workbook

from antares.data_collection.tools.tools import create_xlsx_workbook, edit_xlsx_workbook


##
# Creation
##


def test_create_workbook_dir_not_exist(tmp_path: Path) -> None:
    dir_export_path = tmp_path / "toto"

    # then
    with pytest.raises(
        FileNotFoundError,
        match=re.escape(f"Input directory does not exist: {dir_export_path}"),
    ):
        create_xlsx_workbook(
            path_dir=dir_export_path, workbook_name="dir_not_exist", sheet_name="sheet"
        )


def test_create_empty_workbook(tmp_path: Path) -> None:
    dir_export_path = tmp_path / "links_data_export"
    os.makedirs(dir_export_path)

    # when
    create_xlsx_workbook(
        path_dir=dir_export_path,
        workbook_name="empty_workbook",
        sheet_name="empty_sheet",
    )
    # then
    assert os.path.exists(dir_export_path / "empty_workbook.xlsx")

    wb = load_workbook(filename=dir_export_path / "empty_workbook.xlsx")
    assert wb.sheetnames == ["empty_sheet"]


def test_create_workbook_with_data(tmp_path: Path) -> None:
    dir_export_path = tmp_path / "links_data_export"
    os.makedirs(dir_export_path)

    # given
    df_pandaset = pd.DataFrame(
        {
            "YEAR": [2030, 2040, 2060],
            "STUDY_SCENARIO": ["ERAA", "ERAA", "TYNDP"],
        }
    )

    # when
    create_xlsx_workbook(
        path_dir=dir_export_path,
        workbook_name="test_workbook",
        sheet_name="data_sheet",
        data_df=df_pandaset,
    )

    # then
    df_to_test = pd.read_excel(dir_export_path / "test_workbook.xlsx")

    assert list(df_to_test.columns) == list(df_pandaset.columns)
    assert df_to_test.shape == df_pandaset.shape
    pd.testing.assert_frame_equal(
        df_to_test, df_pandaset, check_dtype=False, check_like=True
    )


def test_create_workbook_overwrite_false(tmp_path: Path) -> None:
    dir_export_path = tmp_path / "links_data_export"
    os.makedirs(dir_export_path)

    # given
    df_pandaset = pd.DataFrame(
        {
            "YEAR": [2030, 2040, 2060],
            "STUDY_SCENARIO": ["ERAA", "ERAA", "TYNDP"],
        }
    )

    # when
    create_xlsx_workbook(
        path_dir=dir_export_path,
        workbook_name="test_workbook_overwrite",
        sheet_name="data_sheet",
        data_df=df_pandaset,
    )

    # then
    name_file = "test_workbook_overwrite" + ".xlsx"
    with pytest.raises(
        FileExistsError,
        match=f"This Workbook already exist: {name_file}",
    ):
        create_xlsx_workbook(
            path_dir=dir_export_path,
            workbook_name="test_workbook_overwrite",
            sheet_name="data_sheet",
            data_df=df_pandaset,
        )


def test_create_workbook_overwrite_true(tmp_path: Path) -> None:
    dir_export_path = tmp_path / "links_data_export"
    os.makedirs(dir_export_path)

    # given
    df_pandaset = pd.DataFrame(
        {
            "YEAR": [2030, 2040, 2060],
            "STUDY_SCENARIO": ["ERAA", "ERAA", "TYNDP"],
        }
    )

    df_pandaset_overwrite = pd.DataFrame(
        {
            "YEAR": [2030, 2040, 2060],
            "STUDY_SCENARIO_NEW": ["ERAA", "ERAA", "TYNDP"],
        }
    )

    # when
    create_xlsx_workbook(
        path_dir=dir_export_path,
        workbook_name="test_workbook_overwrite",
        sheet_name="data_sheet",
        data_df=df_pandaset,
    )

    # overwrite
    create_xlsx_workbook(
        path_dir=dir_export_path,
        workbook_name="test_workbook_overwrite",
        sheet_name="data_sheet_overwrite",
        data_df=df_pandaset_overwrite,
        overwrite=True,
    )

    # then
    wb = load_workbook(filename=dir_export_path / ("test_workbook_overwrite" + ".xlsx"))
    assert wb.sheetnames == ["data_sheet_overwrite"]

    df_to_test = pd.read_excel(dir_export_path / ("test_workbook_overwrite" + ".xlsx"))

    assert list(df_to_test.columns) == list(df_pandaset_overwrite.columns)
    assert df_to_test.shape == df_pandaset_overwrite.shape
    pd.testing.assert_frame_equal(
        df_to_test, df_pandaset_overwrite, check_dtype=False, check_like=True
    )


def test_create_workbook_df_index_header_true(tmp_path: Path) -> None:
    dir_export_path = tmp_path / "links_data_export"
    os.makedirs(dir_export_path)

    # given df + explicit index
    df_pandaset = pd.DataFrame(
        {
            "YEAR": [2030, 2040, 2060],
            "STUDY_SCENARIO": ["ERAA", "ERAA", "TYNDP"],
        },
    )

    # when
    create_xlsx_workbook(
        path_dir=dir_export_path,
        workbook_name="test_workbook_header_index",
        sheet_name="data_sheet",
        data_df=df_pandaset,
        index=True,
        header=True,
    )

    # then
    df_to_test = pd.read_excel(
        dir_export_path / ("test_workbook_header_index" + ".xlsx")
    )
    expected_cols = ["Unnamed: 0", *df_pandaset.columns.to_list()]

    assert df_to_test.columns.to_list() == expected_cols
    rows, cols = df_pandaset.shape
    assert df_to_test.shape == (rows, cols + 1)
    assert all(df_to_test.index.values) == all(df_pandaset.index.values)


##
# Edition
##


def test_edit_workbook_file_not_exist(tmp_path: Path) -> None:
    dir_export_path = tmp_path / "toto.xlsx"

    # then
    with pytest.raises(
        FileNotFoundError,
        match=re.escape(f"This Excel file does not exist: {dir_export_path}"),
    ):
        edit_xlsx_workbook(
            path_file=dir_export_path, sheet_name="sheet", data_df=pd.DataFrame()
        )


def test_edit_workbook_overwrite_false(mock_links_main_params_xlsx: Path) -> None:
    # given
    name_sheet = "STUDY_SCENARIO"
    # then
    with pytest.raises(KeyError, match=f"This sheet already exists: {name_sheet}"):
        edit_xlsx_workbook(
            path_file=mock_links_main_params_xlsx,
            sheet_name=name_sheet,
            data_df=pd.DataFrame(),
        )


def test_edit_workbook_overwrite_sheet_not_exist(
    mock_links_main_params_xlsx: Path,
) -> None:
    # given
    name_sheet = "STUDY_SCENARIOX"
    # then
    with pytest.raises(KeyError, match=f"Sheet '{name_sheet}' not found"):
        edit_xlsx_workbook(
            path_file=mock_links_main_params_xlsx,
            sheet_name=name_sheet,
            data_df=pd.DataFrame(),
            overwrite_sheet=True,
        )


def test_edit_workbook_overwrite_sheet_exist(mock_links_main_params_xlsx: Path) -> None:
    # given
    name_sheet = "STUDY_SCENARIO"
    new_data = pd.DataFrame({"YEAR": [2030, 2040, 2060]})

    # when
    edit_xlsx_workbook(
        path_file=mock_links_main_params_xlsx,
        sheet_name=name_sheet,
        data_df=new_data,
        overwrite_sheet=True,
    )

    # then
    # dtype = {"YEAR": object}
    df_to_test = pd.read_excel(mock_links_main_params_xlsx, sheet_name=name_sheet)

    assert list(new_data.columns) == list(df_to_test.columns)
    assert new_data.shape == df_to_test.shape
    pd.testing.assert_frame_equal(
        df_to_test, new_data, check_dtype=False, check_like=True
    )


def test_edit_workbook_add_new_sheet(mock_links_main_params_xlsx: Path) -> None:
    # given
    name_sheet = "STUDY_SCENARIO_BIS"
    new_data = pd.DataFrame({"YEAR": [2030, 2040, 2060]})

    # when
    edit_xlsx_workbook(
        path_file=mock_links_main_params_xlsx, sheet_name=name_sheet, data_df=new_data
    )

    # then
    df_to_test = pd.read_excel(mock_links_main_params_xlsx, sheet_name=name_sheet)

    assert list(new_data.columns) == list(df_to_test.columns)
    assert df_to_test.shape == df_to_test.shape
    pd.testing.assert_frame_equal(
        df_to_test, new_data, check_dtype=False, check_like=True
    )


def test_edit_workbook_df_index_header_true(mock_links_main_params_xlsx: Path) -> None:
    # given
    name_sheet = "STUDY_SCENARIO_BIS"
    new_data = pd.DataFrame({"YEAR": [2030, 2040, 2060]})

    # when
    edit_xlsx_workbook(
        path_file=mock_links_main_params_xlsx,
        sheet_name=name_sheet,
        data_df=new_data,
        index=True,
        header=True,
    )

    # then
    df_to_test = pd.read_excel(mock_links_main_params_xlsx, sheet_name=name_sheet)
    expected_cols = ["Unnamed: 0", *new_data.columns.to_list()]

    assert df_to_test.columns.to_list() == expected_cols
    rows, cols = new_data.shape
    assert df_to_test.shape == (rows, cols + 1)
    assert all(df_to_test.index.values) == all(new_data.index.values)
