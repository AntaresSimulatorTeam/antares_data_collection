# Copyright (c) 2024, RTE (https://www.rte-france.com)
#
# See AUTHORS.txt
#
# This Source Code Form is subject to the terms of the Mozilla Public
# License, v. 2.0. If a copy of the MPL was not distributed with this
# file, You can obtain one at http://mozilla.org/MPL/2.0/.
#
# SPDX-License-Identifier: MPL-2.0
#
# This file is part of the Antares project.
from dataclasses import dataclass

# structure Referential (MAIN_PARAMS.xlsx)
from enum import Enum
from pathlib import Path

import pandas as pd

from openpyxl.reader.excel import load_workbook


# all workbook sheet names
class ReferentialSheetNames(Enum):
    PAYS = "PAYS"
    STUDY_SCENARIO = "STUDY_SCENARIO"
    LINKS = "LINKS"
    CLUSTER = "CLUSTER"
    PEAK_PARAMS = "PEAK_PARAMS"


# sheet "PAYS"
class CountryColumnsNames(Enum):
    NOM_PAYS = "Nom_pays"
    CODE_PAYS = "code_pays"
    AREAS = "areas"
    MARKET_NODE = "market_node"
    CODE_ANTARES = "code_antares"


# sheet "STUDY_SCENARIO"
class StudyScenarioColumnsNames(Enum):
    YEAR = "YEAR"
    STUDY_SCENARIO = "STUDY_SCENARIO"


# sheet "LINKS"
class LinksColumnsNames(Enum):
    MARKET_NODE = "market_node"
    CODE_ANTARES = "code_antares"


# sheet "CLUSTER"
class ClusterColumnsNames(Enum):
    TYPE = "TYPE"
    CLUSTER_PEMMDB = "CLUSTER_PEMMDB"
    CLUSTER_BP = "CLUSTER_BP"


# "PEAK_PARAMS"
class PeakParamsColumnsNames(Enum):
    HOUR = "hour"
    PERIOD_HOUR = "period_hour"
    MONTH = "month"
    PERIOD_MONTH = "period_month"


@dataclass
class MainParams:
    _market_to_antares: dict[str, str]
    _year_to_scenario: dict[int, str]
    _cluster_pemmdb_to_bp: dict[str, str]

    def get_antares_code(self, market_code: str) -> str:
        if market_code not in self._market_to_antares:
            raise ValueError(f"No antares code defined for market {market_code}")
        return self._market_to_antares[market_code]

    def get_antares_codes(self, market_codes: list[str]) -> list[str]:
        return [self.get_antares_code(c) for c in market_codes]

    def get_scenario_type(self, year: int) -> str:
        if year not in self._year_to_scenario:
            raise ValueError(f"No scenario defined for year {year}")
        return self._year_to_scenario[year]

    def get_scenario_types(self, years: list[int]) -> list[str]:
        return [self.get_scenario_type(y) for y in years]

    def get_cluster_bp(self, cluster_pemmdb: str) -> str:
        if cluster_pemmdb not in self._cluster_pemmdb_to_bp:
            raise ValueError(f"No cluster BP defined for cluster {cluster_pemmdb}")
        return self._cluster_pemmdb_to_bp[cluster_pemmdb]

    def get_clusters_bp(self, clusters_pemmdb: list[str]) -> list[str]:
        return [self.get_cluster_bp(c) for c in clusters_pemmdb]







def parse_main_params(file_path: Path) -> MainParams:
    """
    Parse and validate a MAIN_PARAMS.xlsx workbook.

    This function validates the structure of the provided Excel file:
    - Ensures the file exists.
    - Checks that all required sheets are present.
    - Verifies that each sheet contains exactly the expected columns
      (column order is ignored, but no missing or extra columns are allowed).

    If validation succeeds, the function returns a populated `MainParams`
    dataclass containing the corresponding pandas DataFrames.

    Args:
        file_path (Path): Path to the MAIN_PARAMS.xlsx file to validate.

    Returns:
        MainParams: Dataclass containing validated DataFrames for:
            - pays
            - study_scenario
            - cluster

    Raises:
        FileNotFoundError: If the file does not exist.
        ValueError: If:
            - One or more required sheets are missing.
            - A sheet contains missing or unexpected columns.
    """

    if not file_path.exists():
        raise FileNotFoundError(f"Input file does not exist: {file_path}")

    target_sheet_names = [
        ReferentialSheetNames.PAYS.value,
        ReferentialSheetNames.STUDY_SCENARIO.value,
        ReferentialSheetNames.CLUSTER.value,
        ReferentialSheetNames.PEAK_PARAMS.value,
    ]

    columns_names_dict = {
        ReferentialSheetNames.PAYS.value: [c.value for c in CountryColumnsNames],
        ReferentialSheetNames.STUDY_SCENARIO.value: [c.value for c in StudyScenarioColumnsNames],
        ReferentialSheetNames.CLUSTER.value: [c.value for c in ClusterColumnsNames],
        ReferentialSheetNames.PEAK_PARAMS.value: [c.value for c in PeakParamsColumnsNames],
    }

    # Check sheets
    wb = load_workbook(filename=file_path)
    workbook_sheets = set(wb.sheetnames)

    # missing_sheets = set(target_sheet_names) - workbook_sheets
    missing_sheets = [sheet for sheet in target_sheet_names if sheet not in workbook_sheets]
    if missing_sheets:
        raise ValueError(f"Missing sheets: {missing_sheets}")

    # Read sheets
    dict_of_df = {sheet: pd.read_excel(file_path, sheet_name=sheet) for sheet in target_sheet_names}

    # Validate columns
    for sheet, df in dict_of_df.items():
        expected_cols = set(columns_names_dict[sheet])
        actual_cols = set(df.columns)

        if actual_cols != expected_cols:
            raise ValueError(f"Columns mismatch in sheet '{sheet}'. Expected: {expected_cols}, Got: {actual_cols}")

    # Return validated dataclass
    return MainParams(
        pays=dict_of_df[ReferentialSheetNames.PAYS.value],
        study_scenario=dict_of_df[ReferentialSheetNames.STUDY_SCENARIO.value],
        cluster=dict_of_df[ReferentialSheetNames.CLUSTER.value],
        peak_params=dict_of_df[ReferentialSheetNames.PEAK_PARAMS.value],
    )
