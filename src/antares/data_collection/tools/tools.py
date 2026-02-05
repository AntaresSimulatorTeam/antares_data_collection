# Copyright (c) 2024, RTE (https://www.rte-france.com)
#
# See AUTHORS.txt
#
# This Source Code Form is subject to the terms of the Mozilla Public
# License, v. 2.0. If a copy of the MPL was not distributed with this
# file, You can obtain one at http://mozilla.org/MPL/2.0/.
#
# SPDX-License-Identifier: MPL-2.0
#
# This file is part of the Antares project.
from pathlib import Path
from typing import Optional, List

import pandas as pd
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.workbook.workbook import Workbook
from openpyxl.reader.excel import load_workbook


##
# Data management
##


# TODO add tests
def scenario_filter(
    df_input: pd.DataFrame, filter_params: Optional[List[str]] = None
) -> pd.DataFrame:
    valid_choices: List[str] = ["ERAA", "TYNDP"]

    # default: "ERAA"
    if filter_params is None or len(filter_params) != 1:
        filter_params = ["ERAA"]

    fp: str = filter_params[0]

    if fp not in valid_choices:
        raise ValueError(f"filter_params must be in {valid_choices}")

    filter_map: dict[str, str] = {
        "ERAA": r"ERAA&TYNDP|ERAA",
        "TYNDP": r"ERAA&TYNDP|TYNDP",
    }

    pattern: str = filter_map[fp]

    return df_input[df_input["STUDY_SCENARIO"].str.contains(pattern, regex=True)]


##
# Export : Excel workbook
##


def create_xlsx_workbook(
    path_dir: Path,
    workbook_name: str,
    sheet_name: str,
    data_df: Optional[pd.DataFrame] = None,
    index: bool = False,
    header: bool = True,
    overwrite: bool = False,
) -> None:
    """
    Create an Excel workbook and write a DataFrame to a worksheet.

    Parameters:
        path_dir: Directory where the Excel workbook will be created.
        workbook_name: Workbook name without the `.xlsx` extension.
        sheet_name: Worksheet name to create in the workbook.
        data_df: DataFrame to write into the worksheet. If None, the worksheet
            is created without writing data.
        index: Whether to write the DataFrame index. Defaults to False.
        header: Whether to write the DataFrame column names. Defaults to True.
        overwrite: If True, overwrite the existing Excel workbook. If False
            and the file already exists, raise a FileExistsError.

    Raises:
        FileExistsError: If the Excel workbook already exists and `overwrite` is False.
    """
    if not path_dir.exists():
        raise FileNotFoundError(f"Input directory does not exist: {path_dir}")

    if not overwrite:
        name_file = workbook_name + ".xlsx"
        if (path_dir / name_file).exists():
            raise FileExistsError(f"This Workbook already exist: {name_file}")

    wb = Workbook()
    ws = wb.active
    assert ws is not None  # to prevent error with mypy and "types-openpyxl"
    ws.title = sheet_name
    wb_name = workbook_name + ".xlsx"

    # empty workbook
    if data_df is None:
        wb.save(path_dir / wb_name)

    # workbook with data
    if data_df is not None:
        # --- NORMALISATION PANDAS -> EXCEL ---
        df_excel = data_df.astype(object)
        df_excel = df_excel.mask(pd.isna(df_excel), None)

        for r in dataframe_to_rows(df_excel, index=index, header=header):
            if not all(cell is None for cell in r):
                ws.append(r)

        wb.save(path_dir / wb_name)


def edit_xlsx_workbook(
    path_file: Path,
    sheet_name: str,
    data_df: pd.DataFrame,
    index: bool = False,
    header: bool = True,
    overwrite_sheet: bool = False,
) -> None:
    """
    Edit an existing Excel workbook and write a DataFrame to a new worksheet
    or append to an existing one.

    Parameters:
        path_file: Complete path to the Excel workbook to edit.
        sheet_name: Worksheet name to create in the workbook.
        data_df: DataFrame to write into the worksheet.
        index: Whether to write the DataFrame index. Defaults to False.
        header: Whether to write the DataFrame column names. Defaults to True.
        overwrite_sheet: If True, overwrite the existing worksheet. If False
            and the worksheet already exists, raise a KeyError.

    Raises:
        KeyError: If the worksheet already exists and `overwrite_sheet` is False.
    """
    if not path_file.exists():
        raise FileNotFoundError(f"This Excel file does not exist: {path_file}")

    wb = load_workbook(filename=path_file)
    if not overwrite_sheet:
        if sheet_name in wb.sheetnames:
            raise KeyError(f"This sheet already exists: {sheet_name}")

    if overwrite_sheet:
        if sheet_name not in wb.sheetnames:
            raise KeyError(f"Sheet '{sheet_name}' not found")
        # purge to use append() then
        ws = wb[sheet_name]
        ws.delete_rows(1, ws.max_row)
        wb.save(path_file)

    # add new sheet
    if sheet_name not in wb.sheetnames:
        wb.create_sheet(sheet_name)
        ws = wb[sheet_name]

    # edit workbook in sheet
    if data_df is not None:
        # --- NORMALISATION PANDAS -> EXCEL ---
        df_excel = data_df.astype(object)
        df_excel = df_excel.mask(pd.isna(df_excel), None)
    for r in dataframe_to_rows(df_excel, index=index, header=header):
        if not all(cell is None for cell in r):
            ws.append(r)

    wb.save(path_file)
